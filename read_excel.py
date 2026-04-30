import openpyxl
import json

wb = openpyxl.load_workbook('1.xlsx', read_only=True, data_only=True)
result = {}
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    result[name] = {
        "total_rows": len(rows),
        "header_row": [str(c) if c is not None else "" for c in rows[0]] if rows else [],
        "row_1": [str(c) if c is not None else "" for c in rows[1]] if len(rows) > 1 else [],
    }
wb.close()

with open('excel_headers.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("Done")
