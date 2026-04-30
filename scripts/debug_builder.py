"""Debug DailyReportBuilder với dữ liệu thực từ 1.xlsx"""
import sys
sys.path.insert(0, '/app')

import openpyxl
from app.engines.extraction.daily_report_builder import DailyReportBuilder, _CUSTOM_MAPPING_CACHE

EXCEL_PATH = "/app/1.xlsx"
TEMPLATE_CONFIGS = [
    {
        "worksheet": "BC NGÀY",
        "schema_path": "/app/app/domain/templates/bc_ngay_schema.yaml",
        "target_section": "header",
    },
]

class FakeTemplate:
    google_sheet_configs = TEMPLATE_CONFIGS

def read_excel_sheets(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        if rows:
            sheets[sheet_name] = rows
    return sheets

def main():
    # Clear cache
    _CUSTOM_MAPPING_CACHE.clear()
    
    print("Reading Excel...")
    sheet_data = read_excel_sheets(EXCEL_PATH)
    
    # Chỉ lấy BC NGÀY
    filtered_data = {k: v for k, v in sheet_data.items() if k == "BC NGÀY"}
    print(f"Sheets: {list(filtered_data.keys())}")
    
    # Check rows
    bc_ngay = filtered_data["BC NGÀY"]
    print(f"BC NGÀY rows: {len(bc_ngay)}")
    print(f"Row 0 (header): {bc_ngay[0][:5]}...")
    print(f"Row 3 (data): {bc_ngay[3][:5]}...")
    
    print("\n--- Creating DailyReportBuilder ---")
    builder = DailyReportBuilder(
        template=FakeTemplate(),
        sheet_data=filtered_data,
        worksheet_configs=TEMPLATE_CONFIGS,
    )
    
    print("\n--- Calling build_all_by_date() ---")
    date_reports = builder.build_all_by_date()
    
    print(f"\nResult: {len(date_reports)} reports")
    for dk in sorted(date_reports.keys())[:5]:
        report = date_reports[dk]
        print(f"\n=== {dk} ===")
        print(f"  ngay_bao_cao: '{report.header.ngay_bao_cao}'")
        print(f"  tong_vu_chay: {report.phan_I_va_II_chi_tiet_nghiep_vu.tong_so_vu_chay}")
        print(f"  tong_cnch: {report.phan_I_va_II_chi_tiet_nghiep_vu.tong_so_vu_cnch}")
        print(f"  bang_thong_ke: {len(report.bang_thong_ke)} rows")

if __name__ == "__main__":
    main()
