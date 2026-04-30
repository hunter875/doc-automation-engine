"""Debug DailyReportBuilder với debug chi tiết."""
import sys
sys.path.insert(0, '/app')

import openpyxl
from app.engines.extraction.daily_report_builder import DailyReportBuilder, _normalize_key, _load_custom_mapping

EXCEL_PATH = "/app/1.xlsx"
TEMPLATE_CONFIGS = [
    {
        "worksheet": "BC NGÀY",
        "schema_path": "/app/app/domain/templates/bc_ngay_schema.yaml",
        "target_section": "header",
    },
]

class FakeTemplate:
    google_sheet_configs = TEMPLATE_CONFIGS

def read_excel_sheets(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        if rows:
            sheets[sheet_name] = rows
    return sheets

def main():
    print("Reading Excel...")
    sheet_data = read_excel_sheets(EXCEL_PATH)
    
    # Chỉ lấy BC NGÀY
    filtered_data = {k: v for k, v in sheet_data.items() if k == "BC NGÀY"}
    print(f"Sheets: {list(filtered_data.keys())}")
    
    # Check rows
    bc_ngay = filtered_data["BC NGÀY"]
    print(f"BC NGÀY rows: {len(bc_ngay)}")
    
    # Clear cache
    from app.engines.extraction.daily_report_builder import _CUSTOM_MAPPING_CACHE
    _CUSTOM_MAPPING_CACHE.clear()
    print("Cache cleared")
    
    # Load schema and check mapping
    schema_path = "/app/app/domain/templates/bc_ngay_schema.yaml"
    mapping = _load_custom_mapping(schema_path)
    print(f"\nSchema mapping keys: {list(mapping.keys())[:5]}...")
    
    # Get header
    header_row = bc_ngay[0]
    header_norm = {}
    for idx, h in enumerate(header_row):
        if h is not None:
            header_norm[_normalize_key(str(h))] = idx
    print(f"\nNormalized header: {header_norm}")
    
    # Check header aliases
    header_spec = mapping.get("header", {})
    print(f"\nHeader spec fields: {list(header_spec.keys())}")
    
    # Manual date finding
    day_col = -1
    month_col = -1
    for field_name, spec in header_spec.items():
        aliases = spec.get("aliases", []) if isinstance(spec, dict) else []
        for alias in aliases + [field_name]:
            norm = _normalize_key(str(alias))
            if norm in header_norm:
                if field_name == "ngay_bao_cao_day":
                    day_col = header_norm[norm]
                    print(f"Found day_col: {day_col} (alias: {alias})")
                elif field_name == "ngay_bao_cao_month":
                    month_col = header_norm[norm]
                    print(f"Found month_col: {month_col} (alias: {alias})")
                break
    
    # Check date parsing
    print(f"\nDate columns: day={day_col}, month={month_col}")
    
    if day_col >= 0 and month_col >= 0:
        print("\nDate parsing:")
        for i in range(3, min(8, len(bc_ngay))):
            row = bc_ngay[i]
            day_raw = row[day_col] if day_col < len(row) else None
            month_raw = row[month_col] if month_col < len(row) else None
            print(f"  Row {i}: day={day_raw}, month={month_raw}")
            
            # Try to make date key
            try:
                day_int = int(str(day_raw or "").strip())
                month_int = int(str(month_raw or "").strip())
                if 1 <= day_int <= 31 and 1 <= month_int <= 12:
                    date_key = f"{day_int:02d}/{month_int:02d}"
                    print(f"    -> date_key: {date_key}")
            except:
                print(f"    -> cannot parse date")

if __name__ == "__main__":
    main()
