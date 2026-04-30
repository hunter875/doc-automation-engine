"""Test end-to-end với file Excel thực tế 1.xlsx."""
import sys
sys.path.insert(0, '/app')

import openpyxl

# Clear all schema/mapping caches to pick up latest YAML changes
from app.engines.extraction.sheet_pipeline import _CUSTOM_MAPPING_CACHE
from app.engines.extraction.daily_report_builder import _CUSTOM_MAPPING_CACHE as DRB_CACHE
_CUSTOM_MAPPING_CACHE.clear()
DRB_CACHE.clear()

# Clear the sheet_pipeline _load_sheet_mapping LRU cache too
from app.engines.extraction import sheet_pipeline
if hasattr(sheet_pipeline._load_sheet_mapping, 'cache_clear'):
    sheet_pipeline._load_sheet_mapping.cache_clear()

from app.engines.extraction.daily_report_builder import DailyReportBuilder

EXCEL_PATH = "/app/1.xlsx"
TEMPLATE_CONFIGS = [
    {
        "worksheet": "BC NGÀY",
        "schema_path": "/app/app/domain/templates/bc_ngay_kv30_schema.yaml",
        "target_section": "header",
    },
    {
        "worksheet": "CNCH",
        "schema_path": "/app/app/domain/templates/cnch_kv30_schema.yaml",
        "target_section": "danh_sach_cnch",
    },
    {
        "worksheet": "VỤ CHÁY THỐNG KÊ",
        "schema_path": "/app/app/domain/templates/vu_chay_kv30_schema.yaml",
        "target_section": "danh_sach_chay",
    },
    {
        "worksheet": "CHI VIỆN",
        "schema_path": "/app/app/domain/templates/chi_vien_kv30_schema.yaml",
        "target_section": "danh_sach_chi_vien",
    },
]

class FakeTemplate:
    google_sheet_configs = TEMPLATE_CONFIGS

def read_excel_sheets(path):
    """Đọc tất cả worksheets từ Excel."""
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        if rows:
            sheets[sheet_name] = rows
    return sheets

def main():
    print(f"Đọc Excel: {EXCEL_PATH}")
    try:
        sheet_data = read_excel_sheets(EXCEL_PATH)
        print(f"\nWorksheets tìm thấy: {list(sheet_data.keys())}")
        
        # In thông tin sheets quan trọng
        for name in ["BC NGÀY", "CNCH", "VỤ CHÁY THỐNG KÊ", "CHI VIỆN"]:
            if name in sheet_data:
                print(f"\n=== {name} ({len(sheet_data[name])} rows) ===")
                for i, row in enumerate(sheet_data[name][:3]):
                    print(f"  Row {i}: {row[:10]}...")
    except FileNotFoundError:
        print(f"Không tìm thấy file: {EXCEL_PATH}")
        return

    print("\n" + "="*60)
    print("Chạy ingestion pipeline...")
    
    # Filter chỉ lấy sheets cần thiết
    filtered_data = {
        k: v for k, v in sheet_data.items() 
        if k in ["BC NGÀY", "CNCH", "VỤ CHÁY THỐNG KÊ", "CHI VIỆN"]
    }
    
    print(f"\nSheets sử dụng: {list(filtered_data.keys())}")
    
    # Tạo builder và chạy
    builder = DailyReportBuilder(
        template=FakeTemplate(),
        sheet_data=filtered_data,
        worksheet_configs=TEMPLATE_CONFIGS,
    )
    
    date_reports = builder.build_all_by_date()

    print(f"\nKết quả: {len(date_reports)} báo cáo được tạo")

    for dk in sorted(date_reports.keys())[:10]:  # Chỉ show 10 đầu
        report = date_reports[dk]
        print(f"\n--- {dk} ---")
        print(f"  ngay_bao_cao: {report.header.ngay_bao_cao!r}")
        print(f"  tong_vu_chay: {report.phan_I_va_II_chi_tiet_nghiep_vu.tong_so_vu_chay}")
        print(f"  tong_sclq: {report.phan_I_va_II_chi_tiet_nghiep_vu.tong_sclq}")
        print(f"  tong_chi_vien: {report.phan_I_va_II_chi_tiet_nghiep_vu.tong_chi_vien}")
        print(f"  tong_cnch: {report.phan_I_va_II_chi_tiet_nghiep_vu.tong_so_vu_cnch}")
        print(f"  kien_nghi: {report.phan_I_va_II_chi_tiet_nghiep_vu.chi_tiet_cnch}")
        print(f"  cnch_items: {len(report.danh_sach_cnch)}")
        print(f"  vu_chay_items: {len(report.danh_sach_chay)}")
        print(f"  chi_vien_items: {len(report.danh_sach_chi_vien)}")
        print(f"  bang_thong_ke_rows: {len(report.bang_thong_ke)}")
        if report.bang_thong_ke:
            for item in report.bang_thong_ke[:5]:
                print(f"    BTK: stt={item.stt}, noi_dung={item.noi_dung[:30]!r}, ket_qua={item.ket_qua}")

    if len(date_reports) > 10:
        print(f"\n... và {len(date_reports) - 10} báo cáo khác")

if __name__ == "__main__":
    main()
