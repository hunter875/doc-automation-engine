import openpyxl
import json

wb = openpyxl.load_workbook('1.xlsx', data_only=True)
result = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue

    # Read merged cells info
    merged_ranges = [str(m) for m in ws.merged_cells.ranges]

    # Read first 5 data rows (rows 2-6, index 2-6)
    data_rows = []
    for i in range(2, min(7, len(rows))):
        row_data = {}
        for j, val in enumerate(rows[i]):
            if val is not None:
                row_data[f"col_{j}"] = str(val)
        data_rows.append(row_data)

    result[sheet_name] = {
        "total_rows": len(rows),
        "header_row_0": [str(c) if c is not None else "" for c in rows[0]] if len(rows) > 0 else [],
        "header_row_1": [str(c) if c is not None else "" for c in rows[1]] if len(rows) > 1 else [],
        "data_row_2": data_rows[0] if len(data_rows) > 0 else {},
        "data_row_3": data_rows[1] if len(data_rows) > 1 else {},
        "merged_cells": merged_ranges[:20],
    }

wb.close()

with open('excel_detail.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("Done")
