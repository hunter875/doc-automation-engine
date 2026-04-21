"""Excel KV30 reader: reads THỐNG KÊ CÔNG TÁC NGÀY PC KV30 2026.xlsx.

Converts the multi-sheet Excel workbook into the normalized dict format
expected by SheetExtractionPipeline.normalize().
"""

from __future__ import annotations

import datetime
import io
import re
from pathlib import Path
from typing import Any

import openpyxl

from app.engines.extraction.sheet_pipeline import (
    _to_int,
    _to_text,
)


# ---------------------------------------------------------------------------
# Column index constants for BC NGÀY sheet (34 cols, header rows 1-3)
# Col A=0, B=1, ... AG=32
# ---------------------------------------------------------------------------
_COL_DATE = 0      # A: NGÀY
_COL_MONTH = 1     # B: THÁNG

# Vụ cháy & CNCH
_COL_VU_CHAY = 2          # C: VỤ CHÁY THỐNG KÊ  → STT 2
_COL_SU_CO_LIEN_QUAN = 3  # D: SCLQ PCCC&CNCH       → STT 14
_COL_CHI_VIEN = 4         # E: VỤ CHÁY CHI VIỆN     → tong_chi_vien
_COL_CNCH = 5             # F: CNCH                   → supplement STT 14

# Kiểm tra PCCC
_COL_KT_DINH_KY_N1 = 6   # G: KIỂM TRA ĐỊNH KỲ NHÓM I  → STT 32 (part)
_COL_KT_DINH_KY_N2 = 7   # H: KIỂM TRA ĐỊNH KỲ NHÓM II → STT 32 (add)
_COL_KT_DOT_XUAT_N1 = 8  # I: ĐỘT XUẤT NHÓM I            → STT 33 (part)
_COL_KT_DOT_XUAT_N2 = 9  # J: ĐỘT XUẤT NHÓM II           → STT 33 (add)
_COL_HUONG_DAN = 10       # K: HƯỚNG DẪN                  → STT 31 (tong_co_so)
_COL_KIEN_NGHI = 11       # L: KIẾN NGHỊ
_COL_XU_PHAT = 12         # M: XỬ PHẠT                    → STT 35
_COL_TIEN_PHAT = 13       # N: TIỀN PHẠT (triệu đồng)    → STT 40
_COL_DINH_CHI = 14        # O: ĐÌNH CHỈ                  → STT 37
_COL_PHUC_HOI = 15        # P: PHỤC HỒI

# Tuyên truyền
_COL_TIN_BAI = 16         # Q: TIN BÀI, PHÓNG SỰ         → STT 22 (so_tin_bai)
_COL_SO_LOP_TT = 17       # R: SỐ LỚP TUYÊN TRUYỀN       → STT 26
_COL_SO_NGUOI_TT = 18     # S: SỐ NGƯỜI THAM DỰ          → STT 28
_COL_KHUYEN_CAO = 19      # T: SỐ KHUYẾN CÁO/TỜ RƠI      → STT 29
_COL_SO_LOP_HL = 20      # U: SỐ LỚP HUẤN LUYỆN         → STT 54
_COL_SO_NGUOI_HL = 21     # V: SỐ NGƯỜI THAM DỰ HL       → STT 55

# Tổng
_COL_TONG_LOP = 22        # W: TỔNG SỐ LỚP
_COL_TONG_NGUOI = 23      # X: TỔNG NGƯỜI

# PA PC06 (cơ sở)
_COL_PA06_XD = 24         # Y: PA PC06 XÂY DỰNG          → STT 43
_COL_PA06_TT = 25         # Z: PA PC06 THỰC TẬP          → STT 44

# PA PC08 (CQ Công an)
_COL_PA08_XD = 26         # AA: PA PC08 XÂY DỰNG        → STT 49
_COL_PA08_TT = 27        # AB: PA PC08 THỰC TẬP         → STT 50

# PA PC09 (CNCH CQ CA)
_COL_PA09_XD = 28        # AC: PA PC09 XÂY DỰNG        → STT 52
_COL_PA09_TT = 29        # AD: PA PC09 THỰC TẬP         → STT 53

# PA PC07 (phương tiện giao thông)
_COL_PA07_XD = 30        # AE: PA PC07 XÂY DỰNG        → STT 46
_COL_PA07_TT = 31        # AF: PA PC07 THỰC TẬP         → STT 47

_COL_GHI_CHU = 32        # AG: Ghi chú


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_excel_date(value: Any) -> str:
    """Convert an Excel date value (datetime/date or "dd/mm/yyyy" string) to 'dd/mm/yyyy'."""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, str):
        # Already formatted string
        text = value.strip()
        # Try "dd/mm/yyyy"
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
        if m:
            return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
        return text
    if isinstance(value, (int, float)):
        # Excel serial date
        try:
            dt = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(value))
            return dt.strftime("%d/%m/%Y")
        except Exception:
            return str(value)
    return str(value) if value else ""


def _safe_float_to_int(value: Any) -> int:
    """Convert Excel float/int to int safely (e.g. 1.0 → 1)."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value) if not isinstance(value, bool) else int(value)
    text = str(value).strip()
    if not text:
        return 0
    # Try parsing as float first
    try:
        f = float(text.replace(",", ""))
        return int(f)
    except (ValueError, AttributeError):
        return 0


def _cell(row: list, idx: int, default: Any = None) -> Any:
    """Safely get cell value by 0-based index."""
    if idx < len(row):
        return row[idx]
    return default


def _cell_text(row: list, idx: int) -> str:
    """Get cell value as stripped text."""
    return _to_text(_cell(row, idx))


def _cell_int(row: list, idx: int) -> int:
    """Get cell value as integer."""
    return _safe_float_to_int(_cell(row, idx))


# ---------------------------------------------------------------------------
# ExcelKV30Reader
# ---------------------------------------------------------------------------

class ExcelKV30Reader:
    """Read THỐNG KÊ CÔNG TÁC NGÀY PC KV30 2026.xlsx and return normalized dict.

    Args:
        excel_path: Path to the .xlsx file on disk.
        excel_bytes: Raw bytes of the .xlsx file (alternative to path).
    """

    def __init__(
        self,
        excel_path: Path | str | None = None,
        excel_bytes: bytes | None = None,
    ) -> None:
        if excel_path:
            self._wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        elif excel_bytes:
            self._wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        else:
            raise ValueError("Must provide either excel_path or excel_bytes")

        # Cache parsed data
        self._bc_ngay_rows: list[list] | None = None
        self._cnch_items: list[dict] | None = None
        self._chi_vien_items: list[dict] | None = None
        self._vu_chay_items: list[dict] | None = None

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def get_sheet_names(self) -> list[str]:
        """Return all sheet names in the workbook."""
        return self._wb.sheetnames

    def get_daily_row(self, day: int, month: int | str, year: int = 2026) -> dict[str, Any]:
        """Extract a single daily row from BC NGÀY sheet.

        Args:
            day: Ngày (1-31)
            month: Tháng as int or string (e.g. 4 or "4" or "04")
            year: Năm (default 2026)

        Returns:
            Dict with all numeric fields for that day.
        """
        month_val = int(str(month).strip())
        rows = self._get_bc_ngay_rows()

        for row in rows:
            if len(row) < 6:
                continue
            col_day = _safe_float_to_int(row[_COL_DATE])
            col_month = _safe_float_to_int(row[_COL_MONTH])
            if col_day == day and col_month == month_val:
                return self._row_to_fields(row)
        return {}

    def normalize_for_pipeline(self, day: int, month: int | str, year: int = 2026) -> dict[str, Any]:
        """Build the normalized dict for SheetExtractionPipeline.

        This is the main entry point — returns a dict matching the shape
        expected by SheetExtractionPipeline.normalize().
        """
        daily = self.get_daily_row(day, month, year)
        if not daily:
            raise ValueError(f"No data found for day={day}, month={month}")

        # Build bang_thong_ke rows from BC NGÀY columns
        btk_rows = self._build_bang_thong_ke(daily)

        # Tuyen truyen online (STT 22-25 from BC NGÀY)
        tuyen_online = self._build_tuyen_truyen_online(daily)

        # Header — not stored in BC NGÀY, needs separate input
        header: dict[str, str] = {}

        # Nghiep vu scalars from BC NGÀY
        nghiep_vu = self._build_nghiep_vu(daily)

        return {
            "header": header,
            "nghiep_vu": nghiep_vu,
            "bang_thong_ke": btk_rows,
            "danh_sach_cnch": self.get_cnch_items(),
            "danh_sach_phuong_tien_hu_hong": [],
            "danh_sach_cong_van_tham_muu": [],
            "danh_sach_cong_tac_khac": [],
            "danh_sach_chi_vien": self.get_chi_vien_items(),
            "danh_sach_chay": self.get_vu_chay_items(),
            "tuyen_truyen_online": tuyen_online,
        }

    def get_cnch_items(self) -> list[dict[str, Any]]:
        """Read the CNCH sheet and return list of raw row dicts.

        The mapping to CNCHItem is done by SheetExtractionPipeline.map_to_schema().
        """
        if self._cnch_items is not None:
            return self._cnch_items  # type: ignore[return-value]

        items: list[dict[str, Any]] = []
        sheet_name = self._find_sheet(["CNCH", "cnch"])
        if not sheet_name:
            self._cnch_items = items
            return self._cnch_items  # type: ignore[return-value]

        ws = self._wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Header row is row index 1 (0-based), data starts at row 2
        for row in rows[2:]:
            if not row or not any(v is not None for v in row):
                continue
            raw_date = _cell(row, 2, "")
            ngay_str = _parse_excel_date(raw_date)
            raw_time = _to_text(_cell(row, 3, ""))
            time_str = _normalize_time(raw_time)

            items.append({
                "stt": _cell_int(row, 0),
                "ngay_xay_ra": ngay_str,
                "thoi_gian": time_str,
                "dia_diem": _to_text(_cell(row, 4, "")),
                "noi_dung_tin_bao": _to_text(_cell(row, 1, "")),
                "thiet_hai": _to_text(_cell(row, 7, "")),
                "thong_tin_nan_nhan": _to_text(_cell(row, 8, "")),
            })

        self._cnch_items = items
        return self._cnch_items  # type: ignore[return-value]

    def get_chi_vien_items(self) -> list[dict[str, Any]]:
        """Read the CHI VIỆN sheet and return list of dicts."""
        if self._chi_vien_items is not None:
            return self._chi_vien_items  # type: ignore[return-value]

        items: list[dict[str, Any]] = []
        sheet_name = self._find_sheet(["CHI VIỆN", "CHI VIEN", "CHI_VIEN", "chi_vien", "chi_viện"])
        if not sheet_name:
            self._chi_vien_items = items
            return self._chi_vien_items  # type: ignore[return-value]

        ws = self._wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        for row in rows[2:]:
            if not row or not any(v is not None for v in row):
                continue
            raw_date = _cell(row, 1, "")
            ngay_str = _parse_excel_date(raw_date)
            items.append({
                "stt": _cell_int(row, 0),
                "ngay": ngay_str,
                "dia_diem": _to_text(_cell(row, 2, "")),
                "khu_vuc_quan_ly": _to_text(_cell(row, 3, "")),
                "so_luong_xe": _cell_int(row, 4),
                "thoi_gian_di": _normalize_time(_to_text(_cell(row, 5, ""))),
                "thoi_gian_ve": _normalize_time(_to_text(_cell(row, 6, ""))),
                "chi_huy": _to_text(_cell(row, 7, "")),
                "ghi_chu": _to_text(_cell(row, 8, "")),
            })

        self._chi_vien_items = items
        return self._chi_vien_items  # type: ignore[return-value]

    def get_vu_chay_items(self) -> list[dict[str, Any]]:
        """Read the VỤ CHÁY THỐNG KÊ sheet and return list of dicts."""
        if self._vu_chay_items is not None:
            return self._vu_chay_items  # type: ignore[return-value]

        items: list[dict[str, Any]] = []
        sheet_name = self._find_sheet(["VỤ CHÁY THỐNG KÊ", "VU CHAY", "VU_CHAY", "vu_chay", "vụ cháy thống kê"])
        if not sheet_name:
            self._vu_chay_items = items
            return self._vu_chay_items  # type: ignore[return-value]

        ws = self._wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        for row in rows[2:]:
            if not row or not any(v is not None for v in row):
                continue
            raw_date = _cell(row, 1, "")
            ngay_str = _parse_excel_date(raw_date)
            items.append({
                "stt": _cell_int(row, 0),
                "ngay_xay_ra": ngay_str,
                "thoi_gian": _normalize_time(_to_text(_cell(row, 3, ""))),
                "ten_vu_chay": _to_text(_cell(row, 2, "")),
                "dia_diem": _to_text(_cell(row, 4, "")),
                "nguyen_nhan": _to_text(_cell(row, 6, "")),
                "thiet_hai_nguoi": _to_text(_cell(row, 7, "")),
                "thiet_hai_tai_san": _to_text(_cell(row, 8, "")),
                "thoi_gian_khong_che": _to_text(_cell(row, 11, "")),
                "thoi_gian_dap_tat": _to_text(_cell(row, 12, "")),
                "so_luong_xe": _cell_int(row, 13),
                "chi_huy": _to_text(_cell(row, 14, "")),
            })

        self._vu_chay_items = items
        return self._vu_chay_items  # type: ignore[return-value]

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    def _find_sheet(self, candidates: list[str]) -> str | None:
        """Find first sheet name matching any of the candidates (case-insensitive)."""
        names_lower = {s.lower(): s for s in self._wb.sheetnames}
        for cand in candidates:
            if cand.lower() in names_lower:
                return names_lower[cand.lower()]
        return None

    def _get_bc_ngay_rows(self) -> list[list]:
        """Get all data rows from BC NGÀY sheet (skip header rows 1-3)."""
        if self._bc_ngay_rows is not None:
            return self._bc_ngay_rows

        # Try BC NGÀY first, then BC NGÀY 1 as fallback
        sheet_name = self._find_sheet(["BC NGÀY", "BC_NGAY", "bc_ngay"])
        if not sheet_name:
            sheet_name = self._find_sheet(["BC NGÀY 1", "BC_NGÀY_1", "bc_ngay_1"])
        if not sheet_name:
            self._bc_ngay_rows = []
            return self._bc_ngay_rows

        ws = self._wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        # BC NGÀY has header rows at index 0,1,2 → data starts at index 3
        # But there may be extra blank rows — skip until first non-blank row
        # where col A is a number (the NGÀY column)
        data_rows: list[list] = []
        for row in all_rows[3:]:
            if not row:
                continue
            # Skip rows where first cell is None or text (header text)
            first = _cell(row, 0)
            if first is None:
                continue
            if isinstance(first, str) and not first.strip():
                continue
            data_rows.append(list(row))

        self._bc_ngay_rows = data_rows
        return self._bc_ngay_rows

    def _row_to_fields(self, row: list) -> dict[str, Any]:
        """Convert a BC NGÀY row to a flat field dict."""
        return {
            "ngay": _cell_int(row, _COL_DATE),
            "thang": _cell_int(row, _COL_MONTH),
            "vu_chay": _cell_int(row, _COL_VU_CHAY),
            "su_co_lien_quan": _cell_int(row, _COL_SU_CO_LIEN_QUAN),
            "chi_vien": _cell_int(row, _COL_CHI_VIEN),
            "cnch": _cell_int(row, _COL_CNCH),
            "kt_dinh_ky_n1": _cell_int(row, _COL_KT_DINH_KY_N1),
            "kt_dinh_ky_n2": _cell_int(row, _COL_KT_DINH_KY_N2),
            "kt_dot_xuat_n1": _cell_int(row, _COL_KT_DOT_XUAT_N1),
            "kt_dot_xuat_n2": _cell_int(row, _COL_KT_DOT_XUAT_N2),
            "huong_dan": _cell_int(row, _COL_HUONG_DAN),
            "kien_nghi": _cell_int(row, _COL_KIEN_NGHI),
            "xu_phat": _cell_int(row, _COL_XU_PHAT),
            "tien_phat": _cell_int(row, _COL_TIEN_PHAT),
            "dinh_chi": _cell_int(row, _COL_DINH_CHI),
            "phuc_hoi": _cell_int(row, _COL_PHUC_HOI),
            "tin_bai": _cell_int(row, _COL_TIN_BAI),
            "so_lop_tt": _cell_int(row, _COL_SO_LOP_TT),
            "so_nguoi_tt": _cell_int(row, _COL_SO_NGUOI_TT),
            "khuyen_cao": _cell_int(row, _COL_KHUYEN_CAO),
            "so_lop_hl": _cell_int(row, _COL_SO_LOP_HL),
            "so_nguoi_hl": _cell_int(row, _COL_SO_NGUOI_HL),
            "tong_lop": _cell_int(row, _COL_TONG_LOP),
            "tong_nguoi": _cell_int(row, _COL_TONG_NGUOI),
            "pa06_xd": _cell_int(row, _COL_PA06_XD),
            "pa06_tt": _cell_int(row, _COL_PA06_TT),
            "pa08_xd": _cell_int(row, _COL_PA08_XD),
            "pa08_tt": _cell_int(row, _COL_PA08_TT),
            "pa09_xd": _cell_int(row, _COL_PA09_XD),
            "pa09_tt": _cell_int(row, _COL_PA09_TT),
            "pa07_xd": _cell_int(row, _COL_PA07_XD),
            "pa07_tt": _cell_int(row, _COL_PA07_TT),
            "ghi_chu": _cell_text(row, _COL_GHI_CHU),
        }

    def _build_bang_thong_ke(self, d: dict[str, Any]) -> list[dict[str, Any]]:
        """Build bang_thong_ke list of dicts from BC NGÀY daily fields."""
        rows: list[dict[str, Any]] = []

        def add(stt: str, noi_dung: str, ket_qua: int) -> None:
            if ket_qua != 0 or noi_dung:
                rows.append({"stt": stt, "noi_dung": noi_dung, "ket_qua": ket_qua})

        # I. Tình hình cháy, nổ
        add("1", "I. TÌNH HÌNH CHÁY, NỔ, SỰ CỐ TAI NẠN", 0)

        # Vụ cháy
        add("2", "1. Tổng số vụ cháy", d.get("vu_chay", 0))

        # Vụ nổ (= vụ chi viện — column E)
        add("8", "2. Tổng số vụ nổ", d.get("chi_vien", 0))

        # CNCH / Sự cố liên quan
        cnch_val = d.get("cnch", 0) or d.get("su_co_lien_quan", 0)
        add("14", "3. Tổng số vụ tai nạn, sự cố", cnch_val)

        add("20", "II. KẾT QUẢ CÔNG TÁC PCCC VÀ CNCH", 0)

        # 1. Tuyên truyền
        add("21", "1. Tuyên truyền về PCCC và CNCH", 0)

        # 1.1 Tuyên truyền online (from col Q = TIN BÀI)
        tin_bai = d.get("tin_bai", 0)
        add("22", "1.1 Tuyên truyền qua các phương tiện thông tin truyền thông và nền tảng trực tuyến MXH", tin_bai)

        # Số tin bài = STT 22
        add("23", "Số tin, bài đã đăng phát", tin_bai)

        # Số hình ảnh — not available in BC NGÀY, set 0
        add("24", "Số hình ảnh được đăng tải", 0)

        # 1.2 Tuyên truyền trực tiếp
        add("26", "1.2 Tuyên truyền trực tiếp tại cơ sở, doanh nghiệp, các khu dân cư", 0)
        add("27", "Số cuộc", d.get("so_lop_tt", 0))
        add("28", "Số người tham dự", d.get("so_nguoi_tt", 0))
        add("29", "Số khuyến cáo, tờ rơi đã phát hành", d.get("khuyen_cao", 0))

        # 2. Kiểm tra
        add("30", "2. Hướng dẫn, kiểm tra về PCCC và CNCH", 0)
        kd_n1 = d.get("kt_dinh_ky_n1", 0)
        kd_n2 = d.get("kt_dinh_ky_n2", 0)
        kd_total = kd_n1 + kd_n2
        dx_n1 = d.get("kt_dot_xuat_n1", 0)
        dx_n2 = d.get("kt_dot_xuat_n2", 0)
        dx_total = dx_n1 + dx_n2
        tong_kiem_tra = d.get("huong_dan", 0) or kd_total + dx_total
        add("31", f"Số cơ sở được kiểm an toàn PCCC (=STT 31+STT 33)", tong_kiem_tra)
        add("32", "Kiểm tra định kỳ", kd_total)
        add("33", "Kiểm tra đột xuất theo chuyên đề", dx_total)
        add("34", "Số vi phạm được phát hiện", d.get("kien_nghi", 0))

        xu_phat = d.get("xu_phat", 0)
        dinh_chi = d.get("dinh_chi", 0)
        phuc_hoi = d.get("phuc_hoi", 0)
        add("35", f"Tổng số cơ sở bị xử phạt VPHC về PCCC (=STT 36+…+STT 39)", xu_phat)
        add("36", "Trong đó, phạt cảnh cáo", 0)
        add("37", "Trong đó, tạm đình chỉ hoạt động", dinh_chi)
        add("38", "Trong đó, đình chỉ hoạt động", 0)
        add("39", "Trong đó, phạt tiền", xu_phat - dinh_chi - phuc_hoi)
        add("40", "Số tiền phạt thu được (triệu đồng)", d.get("tien_phat", 0))

        # 3. Phương án
        add("41", "3. Xây dựng, thực tập phương án chữa cháy, CNCH", 0)
        add("42", "3.1 Cơ sở theo Mẫu số PC06", 0)
        add("43", "Số phương án được xây dựng và phê duyệt", d.get("pa06_xd", 0))
        add("44", "Số phương án được thực tập", d.get("pa06_tt", 0))
        add("45", "3.2 Phương tiện giao thông theo Mẫu số PC07", 0)
        add("46", "Số phương án được xây dựng và phê duyệt", d.get("pa07_xd", 0))
        add("47", "Số phương án được thực tập", d.get("pa07_tt", 0))
        add("48", "3.3 CQ Công an thực hiện theo Mẫu số PC08", 0)
        add("49", "Số phương án được xây dựng và phê duyệt", d.get("pa08_xd", 0))
        add("50", "Số phương án được thực tập", d.get("pa08_tt", 0))
        add("51", "3.4 CNCH CQ Công an thực hiện theo Mẫu số PC09", 0)
        add("52", "Số phương án được xây dựng và phê duyệt", d.get("pa09_xd", 0))
        add("53", "Số phương án được thực tập", d.get("pa09_tt", 0))

        # 4. Huấn luyện
        add("54", "4. Công tác huấn luyện nghiệp vụ chữa cháy và CNCH thường xuyên", 0)
        add("55", "Tổng số CBCS tham gia huấn luyện (=STT 56+…+STT 61)", d.get("so_nguoi_hl", 0))
        add("56", "Chỉ huy phòng", 0)
        add("57", "Chỉ huy Đội", 0)
        add("58", "Cán bộ tiểu đội", 0)
        add("59", "Chiến sĩ CC và CNCH", 0)
        add("61", "Lái tàu CC và CNCH", 0)

        return rows

    def _build_nghiep_vu(self, d: dict[str, Any]) -> dict[str, Any]:
        """Build nghiep_vu dict from BC NGÀY daily fields."""
        return {
            "tong_so_vu_chay": d.get("vu_chay", 0),
            "tong_so_vu_no": d.get("chi_vien", 0),
            "tong_so_vu_cnch": d.get("cnch", 0) or d.get("su_co_lien_quan", 0),
            "tong_chi_vien": d.get("chi_vien", 0),
            "tong_xe_hu_hong": 0,  # read from narrative, not in BC NGÀY
            "tong_tin_bai": d.get("tin_bai", 0),
            "tong_hinh_anh": 0,
            "so_lan_cai_app_114": 0,
        }

    def _build_tuyen_truyen_online(self, d: dict[str, Any]) -> dict[str, int]:
        """Build tuyen_truyen_online dict from BC NGÀY daily fields."""
        return {
            "so_tin_bai": d.get("tin_bai", 0),
            "so_hinh_anh": 0,
            "cai_app_114": 0,
        }


# ---------------------------------------------------------------------------
# Time normalization
# ---------------------------------------------------------------------------

def _normalize_time(raw: str) -> str:
    """Normalize Vietnamese time strings to 'HH:MM ngày dd/mm/yyyy' or 'HH:MM'."""
    raw = (raw or "").strip()
    if not raw:
        return ""

    # Already formatted: "15:10 ngày 31/03/2026" or "15:10"
    m = re.match(r"(\d{1,2}):(\d{2})(?:\s+ngày\s+(\d{1,2})/(\d{1,2})/(\d{4}))?", raw)
    if m:
        hh = int(m.group(1))
        mm = m.group(2)
        if m.group(3):
            return f"{hh:02d}:{mm} ngày {int(m.group(3)):02d}/{int(m.group(4)):02d}/{m.group(5)}"
        return f"{hh:02d}:{mm}"

    # "21 giờ 30 phút"
    m2 = re.match(r"(\d{1,2})\s*giờ?\s*(\d{1,2})?\s*phút?", raw, re.IGNORECASE)
    if m2:
        hh = int(m2.group(1))
        mm = m2.group(2) or "00"
        return f"{hh:02d}:{mm}"

    return raw
