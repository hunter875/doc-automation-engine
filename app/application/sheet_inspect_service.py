"""Sheet Inspector — business logic service.

Provides per-job / per-day inspection of extracted STT data,
comparison against raw Excel bytes from MinIO, and column→STT mapping.
"""

from __future__ import annotations

import io
import logging
import uuid
from datetime import date
from pathlib import Path
from typing import Any, Optional

import openpyxl
import yaml
from sqlalchemy.orm import Session

from app.core.config import settings
from app.domain.models.extraction_job import ExtractionJob
from app.domain.models.document import Document

logger = logging.getLogger(__name__)

# ── STT definitions ─────────────────────────────────────────────────────────────────

# STT 1–61 only — sourced from bc_ngay_schema.yaml bang_thong_ke.stt_map
# Section-header STTs (is_section=True) are skipped in issue detection.
STT_DEFINITIONS: dict[str, dict] = {
    "1":  {"label": "I. TÌNH HÌNH CHÁY, NỔ, SỰ CỐ TAI NẠN", "is_section": True},
    "2":  {"label": "1. Tổng số vụ cháy",                       "is_section": False},
    "3":  {"label": "Số người chết",                            "is_section": False},
    "4":  {"label": "Số người bị thương",                       "is_section": False},
    "5":  {"label": "Số người cứu được",                        "is_section": False},
    "6":  {"label": "Tài sản thiệt hại",                        "is_section": False},
    "7":  {"label": "Tài sản cứu được",                          "is_section": False},
    "8":  {"label": "2. Tổng số vụ nổ",                         "is_section": False},
    "9":  {"label": "Số người chết",                             "is_section": False},
    "10": {"label": "Số người bị thương",                       "is_section": False},
    "11": {"label": "Số người cứu được",                         "is_section": False},
    "12": {"label": "Tài sản thiệt hại",                         "is_section": False},
    "13": {"label": "Tài sản cứu được",                          "is_section": False},
    "14": {"label": "3. Tổng số vụ tai nạn, sự cố",            "is_section": False},
    "15": {"label": "Số người cứu được (=16+17)",               "is_section": False},
    "16": {"label": "Trực tiếp cứu được",                        "is_section": False},
    "17": {"label": "Hướng dẫn thoát nạn",                      "is_section": False},
    "18": {"label": "Số thi thể",                                 "is_section": False},
    "19": {"label": "Tài sản cứu được",                          "is_section": False},
    "20": {"label": "II. KẾT QUẢ CÔNG TÁC PCCC VÀ CNCH",        "is_section": True},
    "21": {"label": "1. Tuyên truyền về PCCC và CNCH",           "is_section": True},
    "22": {"label": "1.1 Tuyên truyền qua các phương tiện thông tin truyền thông và nền tảng trực tuyến MXH", "is_section": False},
    "23": {"label": "Số tin, bài đã đăng phát",                 "is_section": False},
    "24": {"label": "Số hình ảnh được đăng tải",                 "is_section": False},
    "25": {"label": "Số lượt cài đặt ứng dụng HELP 114",       "is_section": False},
    "26": {"label": "1.2 Tuyên truyền trực tiếp tại cơ sở, doanh nghiệp, các khu dân cư", "is_section": True},
    "27": {"label": "Số cuộc",                                   "is_section": False},
    "28": {"label": "Số người tham dự",                           "is_section": False},
    "29": {"label": "Số khuyến cáo, tờ rơi đã phát hành",       "is_section": False},
    "30": {"label": "2. Hướng dẫn, kiểm tra về PCCC và CNCH",   "is_section": True},
    "31": {"label": "Số cơ sở được kiểm an toàn PCCC (=STT 32+STT 33)", "is_section": False},
    "32": {"label": "Kiểm tra định kỳ",                          "is_section": False},
    "33": {"label": "Kiểm tra đột xuất theo chuyên đề",          "is_section": False},
    "34": {"label": "Số vi phạm được phát hiện",                  "is_section": False},
    "35": {"label": "Tổng số cơ sở bị xử phạt VPHC về PCCC (=STT 36+…+STT 39)", "is_section": False},
    "36": {"label": "Trong đó, phạt cảnh cáo",                   "is_section": False},
    "37": {"label": "Trong đó, tạm đình chỉ hoạt động",          "is_section": False},
    "38": {"label": "Trong đó, đình chỉ hoạt động",              "is_section": False},
    "39": {"label": "Trong đó, phạt tiền",                        "is_section": False},
    "40": {"label": "Số tiền phạt thu được (triệu đồng)",        "is_section": False},
    "41": {"label": "3. Xây dựng, thực tập phương án chữa cháy, CNCH", "is_section": True},
    "42": {"label": "3.1 Cơ sở theo Mẫu số PC06",               "is_section": True},
    "43": {"label": "Số phương án được xây dựng và phê duyệt",  "is_section": False},
    "44": {"label": "Số phương án được thực tập",                 "is_section": False},
    "45": {"label": "3.2 Phương tiện giao thông theo Mẫu số PC07", "is_section": True},
    "46": {"label": "Số phương án được xây dựng và phê duyệt",  "is_section": False},
    "47": {"label": "Số phương án được thực tập",                 "is_section": False},
    "48": {"label": "3.3 CQ Công an thực hiện theo Mẫu số PC08", "is_section": True},
    "49": {"label": "Số phương án được xây dựng và phê duyệt",   "is_section": False},
    "50": {"label": "Số phương án được thực tập",                 "is_section": False},
    "51": {"label": "3.4 CNCH CQ Công an thực hiện theo Mẫu số PC09", "is_section": True},
    "52": {"label": "Số phương án được xây dựng và phê duyệt",   "is_section": False},
    "53": {"label": "Số phương án được thực tập",                 "is_section": False},
    "54": {"label": "4. Công tác huấn luyện nghiệp vụ chữa cháy và CNCH thường xuyên", "is_section": True},
    "55": {"label": "Tổng số CBCS tham gia huấn luyện (=STT 56+…+STT 61)", "is_section": False},
    "56": {"label": "Chỉ huy phòng",                              "is_section": False},
    "57": {"label": "Chỉ huy Đội",                                "is_section": False},
    "58": {"label": "Cán bộ tiểu đội",                            "is_section": False},
    "59": {"label": "Chiến sĩ CC và CNCH",                        "is_section": False},
    "60": {"label": "Chiến sĩ nghĩa vụ (hợp đồng lao động)",    "is_section": False},
    "61": {"label": "Lái tàu CC và CNCH",                         "is_section": False},
}

# Section-header STTs that should never be flagged as zero-count issues
_SECTION_HEADERS = {"1", "20", "21", "26", "30", "41", "42", "45", "48", "51", "54"}

# Key STTs to show in the grid tab (most important ones)
GRID_STTS = ["2", "14", "22", "31", "32", "33", "43", "47", "50", "52", "55", "60", "61"]

# ── YAML mapping ─────────────────────────────────────────────────────────────────

def _load_sheet_mapping() -> dict[str, Any]:
    """Load the sheet_mapping.yaml as a plain dict."""
    yaml_path = Path(__file__).resolve().parent.parent / "domain" / "templates" / "sheet_mapping.yaml"
    if not yaml_path.exists():
        logger.warning("sheet_mapping.yaml not found at %s", yaml_path)
        return {}
    with open(yaml_path, encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


# ── S3 helpers ─────────────────────────────────────────────────────────────────

def _get_excel_bytes(doc: Document) -> Optional[bytes]:
    """Download Excel file from MinIO by document.s3_key."""
    if not doc.s3_key:
        return None
    try:
        from app.application.doc_service import s3_client
        resp = s3_client.get_object(
            Bucket=settings.S3_BUCKET_NAME,
            Key=doc.s3_key,
        )
        return resp["Body"].read()
    except Exception as e:
        logger.warning("Could not download Excel from S3 key=%s: %s", doc.s3_key, e)
        return None


# ── Row helpers ─────────────────────────────────────────────────────────────────

def _safe_float(v: Any) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _safe_int(v: Any) -> int:
    return int(_safe_float(v))


def _as_list(v: Any) -> list:
    if isinstance(v, list):
        return v
    return []


def _as_dict(v: Any) -> dict:
    if isinstance(v, dict):
        return v
    return {}


def _stt_coverage(btk: list[dict]) -> dict[str, int]:
    """Compute STT coverage from bang_thong_ke rows."""
    present = {str(r.get("stt", "")).strip(): r.get("ket_qua", 0) for r in btk if r.get("stt")}
    total = 61
    populated = sum(1 for v in present.values() if v != 0)
    zero = sum(1 for v in present.values() if v == 0)
    missing = total - len(present)
    return {
        "total": total,
        "populated": populated,
        "zero": zero,
        "missing": missing,
    }


def _named_list_counts(data: dict) -> dict[str, int]:
    """Count items in each named list."""
    return {
        "cnch":        len(_as_list(data.get("danh_sach_cnch"))),
        "chay":        len(_as_list(data.get("danh_sach_chay"))),
        "chi_vien":    len(_as_list(data.get("danh_sach_chi_vien"))),
        "phuong_tien": len(_as_list(data.get("danh_sach_phuong_tien_hu_hong"))),
        "cong_van":    len(_as_list(data.get("danh_sach_cong_van_tham_muu"))),
    }


def _extract_stt_values(data: dict) -> dict[str, int]:
    """Extract flat STT values from a job's extracted/enriched/reviewed data."""
    btk = _as_list(data.get("bang_thong_ke"))
    values: dict[str, int] = {}
    for row in btk:
        stt = str(row.get("stt", "")).strip()
        if stt:
            values[f"stt_{int(stt):02d}"] = _safe_int(row.get("ket_qua", 0))
    return values


# ── Main service ────────────────────────────────────────────────────────────────

class SheetInspectService:
    def __init__(self, db: Session) -> None:
        self.db = db

    # ── 1. Month / day inspection ──────────────────────────────────────────────

    def get_month_data(
        self,
        tenant_id: uuid.UUID,
        month: int,
        year: int,
        day: Optional[int] = None,
    ) -> list[dict]:
        """Return all days in the month, each with job-level STT coverage."""
        from calendar import monthrange
        from datetime import timedelta

        _, last_day = monthrange(year, month)
        start = date(year, month, 1)
        end = date(year, month, last_day)

        query = self.db.query(ExtractionJob).filter(
            ExtractionJob.tenant_id == tenant_id,
            ExtractionJob.created_at >= start,
            ExtractionJob.created_at <= end,
        )
        if day is not None:
            from datetime import timedelta as _td
            day_start = date(year, month, day)
            day_end = day_start + _td(days=1)
            query = query.filter(
                ExtractionJob.created_at >= day_start,
                ExtractionJob.created_at < day_end,
            )

        jobs = query.order_by(ExtractionJob.created_at.desc()).all()

        # Group by date
        from collections import defaultdict
        by_date: dict[str, dict] = defaultdict(lambda: {
            "date": "", "job_count": 0, "approved_count": 0,
            "jobs": [], "issue_count": 0,
        })
        for job in jobs:
            day_str = (job.created_at.date() if job.created_at else start).isoformat()
            entry = by_date[day_str]
            entry["date"] = day_str
            entry["job_count"] += 1
            if job.status in ("approved", "aggregated"):
                entry["approved_count"] += 1

            # Best data source
            raw = job.reviewed_data or job.extracted_data or {}
            btk = _as_list(raw.get("bang_thong_ke"))

            stt_vals = _extract_stt_values(raw)
            job_issue_count = 0
            if len(btk) < 20:
                job_issue_count += 1
            present_stts = {str(r.get("stt", "")).strip() for r in btk if r.get("stt")}
            for stt_num, defn in STT_DEFINITIONS.items():
                if defn["is_section"]:
                    continue
                if stt_num not in present_stts:
                    job_issue_count += 1
                elif stt_vals.get(f"stt_{int(stt_num):02d}", 0) == 0:
                    job_issue_count += 1
            entry["issue_count"] += job_issue_count

            entry["jobs"].append({
                "id": str(job.id),
                "file_name": job.file_name or "",
                "status": job.status or "",
                "parser_used": job.parser_used or "",
                "created_at": job.created_at.isoformat() if job.created_at else "",
                "stt_coverage": _stt_coverage(btk),
                "stt_values": stt_vals,
                "btk_rows": btk,
                "named_lists": _named_list_counts(raw),
                "issue_count": job_issue_count,
            })

        # Fill all days
        result = []
        current = start
        while current <= end:
            ds = current.isoformat()
            if ds in by_date:
                result.append(by_date[ds])
            else:
                result.append({"date": ds, "job_count": 0, "approved_count": 0, "jobs": [], "issue_count": 0})
            current += timedelta(days=1)
        return result

    # ── 2. Issues ─────────────────────────────────────────────────────────────

    def get_issues(
        self,
        tenant_id: uuid.UUID,
        month: int,
        year: int,
        document_id: Optional[uuid.UUID] = None,
        job_id: Optional[uuid.UUID] = None,
        worksheet: Optional[str] = None,
    ) -> list[dict]:
        """Identify STT fields that are zero or missing in extracted data,
        optionally comparing against the raw Excel from MinIO.

        Args:
            worksheet: if provided, only check jobs where parser_used matches.
                       If "BC NGÀY" → check bang_thong_ke STT fields.
                       If "VỤ CHÁY THỐNG KÊ" → check danh_sach_chay.
                       If "CNCH" → check danh_sach_cnch.
                       If "CHI VIỆN" → check danh_sach_chi_vien.
                       If None → check all worksheets (legacy behaviour).
        """
        issues: list[dict] = []

        # Load jobs for the month
        days = self.get_month_data(tenant_id, month, year)
        for day_data in days:
            for job_data in day_data.get("jobs", []):
                if job_id and str(job_data["id"]) != str(job_id):
                    continue

                parser = job_data.get("parser_used", "") or ""
                # Filter by worksheet if specified
                if worksheet and worksheet not in parser:
                    continue

                worksheet_issues = self._build_worksheet_issues(
                    job_data, day_data["date"], worksheet
                )
                issues.extend(worksheet_issues)

                # If document_id given, also fetch raw Excel and compare
                # (only makes sense for BC NGÀY worksheet)
                if document_id and (worksheet == "BC NGÀY" or (not worksheet and "BC NGÀY" in parser)):
                    doc = self.db.query(Document).filter(Document.id == document_id).first()
                    if doc:
                        excel_bytes = _get_excel_bytes(doc)
                        if excel_bytes:
                            btk = _as_list(job_data.get("btk_rows", []))
                            issues.extend(self._compare_with_excel(
                                excel_bytes, btk, day_data["date"],
                                job_data["id"], job_data["file_name"], parser
                            ))

        # Deduplicate by stable key: (job_id, worksheet, stt_or_row_idx, issue_type)
        seen: set[tuple] = set()
        deduped: list[dict] = []
        for issue in issues:
            key = (
                issue["job_id"],
                issue.get("worksheet", ""),
                issue.get("stt", ""),
                issue.get("row_index", ""),
                issue["severity"],
            )
            if key not in seen:
                seen.add(key)
                deduped.append(issue)

        return deduped

    def _build_worksheet_issues(
        self,
        job_data: dict,
        date_str: str,
        worksheet: Optional[str],
    ) -> list[dict]:
        """Build issues for a single job based on its worksheet type."""
        issues: list[dict] = []
        raw = job_data  # already a dict from get_month_data output
        parser = raw.get("parser_used", "") or ""

        if worksheet == "BC NGÀY" or (not worksheet and "BC NGÀY" in parser):
            # BC NGÀY: check bang_thong_ke STT fields (STT 1–61)
            btk = _as_list(raw.get("btk_rows", []))
            present_stts = {str(r.get("stt", "")).strip(): r.get("ket_qua", 0) for r in btk if r.get("stt")}

            for stt_num, defn in STT_DEFINITIONS.items():
                if defn["is_section"]:
                    continue
                stt_val = present_stts.get(stt_num)
                stt_key = f"stt_{int(stt_num):02d}"
                if stt_val is None:
                    issues.append({
                        "stt": stt_num,
                        "field": stt_key,
                        "label": defn["label"],
                        "date": date_str,
                        "job_id": job_data["id"],
                        "file_name": job_data["file_name"],
                        "worksheet": "BC NGÀY",
                        "severity": "missing",
                        "excel_value": None,
                        "system_value": 0,
                        "description": f"STT {stt_num} không có trong dữ liệu trích xuất",
                    })
                elif stt_val == 0 and stt_num not in _SECTION_HEADERS:
                    issues.append({
                        "stt": stt_num,
                        "field": stt_key,
                        "label": defn["label"],
                        "date": date_str,
                        "job_id": job_data["id"],
                        "file_name": job_data["file_name"],
                        "worksheet": "BC NGÀY",
                        "severity": "zero",
                        "excel_value": 0,
                        "system_value": 0,
                        "description": f"STT {stt_num} có giá trị = 0",
                    })

        elif worksheet == "VỤ CHÁY THỐNG KÊ" or "CHÁY" in parser.upper():
            chay_rows = _as_list(raw.get("btk_rows", []))  # btk_rows holds raw event rows for event sheets
            if len(chay_rows) == 0:
                issues.append({
                    "stt": "1",
                    "field": "danh_sach_chay",
                    "label": "Danh sách vụ cháy",
                    "date": date_str,
                    "job_id": job_data["id"],
                    "file_name": job_data["file_name"],
                    "worksheet": "VỤ CHÁY THỐNG KÊ",
                    "severity": "missing",
                    "excel_value": None,
                    "system_value": 0,
                    "description": "Sheet VỤ CHÁY THỐNG KÊ không có dữ liệu trong dữ liệu trích xuất",
                })

        elif worksheet == "CNCH" or "CNCH" in parser.upper():
            cnch_rows = _as_list(raw.get("btk_rows", []))
            if len(cnch_rows) == 0:
                issues.append({
                    "stt": "1",
                    "field": "danh_sach_cnch",
                    "label": "Danh sách CNCH",
                    "date": date_str,
                    "job_id": job_data["id"],
                    "file_name": job_data["file_name"],
                    "worksheet": "CNCH",
                    "severity": "missing",
                    "excel_value": None,
                    "system_value": 0,
                    "description": "Sheet CNCH không có dữ liệu trong dữ liệu trích xuất",
                })

        elif worksheet == "CHI VIỆN" or "CHI VIỆN" in parser.upper():
            cv_rows = _as_list(raw.get("btk_rows", []))
            if len(cv_rows) == 0:
                issues.append({
                    "stt": "1",
                    "field": "danh_sach_chi_vien",
                    "label": "Danh sách chi viện",
                    "date": date_str,
                    "job_id": job_data["id"],
                    "file_name": job_data["file_name"],
                    "worksheet": "CHI VIỆN",
                    "severity": "missing",
                    "excel_value": None,
                    "system_value": 0,
                    "description": "Sheet CHI VIỆN không có dữ liệu trong dữ liệu trích xuất",
                })

        return issues

    def _compare_with_excel(
        self,
        excel_bytes: bytes,
        btk: list[dict],
        date_str: str,
        job_id: str,
        file_name: str,
        parser: str = "",
    ) -> list[dict]:
        """Read raw Excel and compare ket_qua values with extracted data."""
        issues: list[dict] = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            if "BC NGÀY" not in wb.sheetnames:
                return issues

            ws = wb["BC NGÀY"]
            # BC NGÀY col layout (0-indexed):
            # col 0=NGÀY, col 2=VỤ CHÁY→STT2, col 5=CNCH→STT14,
            # col 16=TIN BÀI→STT22, col 6=KIỂM TRA DK→STT32
            col_stt_map = {
                2: "stt_02",   # VỤ CHÁY THỐNG KÊ
                5: "stt_14",   # CNCH
                6: "stt_31",   # TỔNG KIỂM TRA
                16: "stt_22",  # TIN BÀI
                20: "stt_55",  # TỔNG CBCS HUẤN LUYỆN
            }
            btk_dict = {str(r.get("stt", "")): r.get("ket_qua", 0) for r in btk if r.get("stt")}

            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or len(row) < 20:
                    continue
                day_val = _safe_int(row[0])
                month_val = _safe_int(row[1])
                target_date = f"{int(month_val):02d}/{int(day_val):02d}/{date_str.split('-')[0]}"
                # Only compare rows matching the date
                if str(day_val) != date_str.split("-")[2].lstrip("0"):
                    continue
                for col_idx, stt_key in col_stt_map.items():
                    excel_val = _safe_int(row[col_idx] if len(row) > col_idx else 0)
                    sys_val = btk_dict.get(stt_key, 0)
                    if sys_val != excel_val and excel_val != 0:
                        stt_num = stt_key.split("_")[1].lstrip("0")
                        defn = STT_DEFINITIONS.get(stt_num, {})
                        label = defn.get("label", stt_key) if defn else stt_key
                        issues.append({
                            "stt": stt_num,
                            "field": stt_key,
                            "label": label,
                            "date": date_str,
                            "job_id": job_id,
                            "file_name": file_name,
                            "worksheet": "BC NGÀY",
                            "severity": "mismatch",
                            "excel_value": excel_val,
                            "system_value": sys_val,
                            "description": f"Mismatch: Excel={excel_val}, Hệ thống={sys_val}",
                        })
                break  # Only first matching row
        except Exception as e:
            logger.warning("Could not read Excel for comparison: %s", e)
        return issues

    # ── 3. Column → STT mapping ─────────────────────────────────────────────

    def get_mapping(self) -> list[dict]:
        """Return column → STT mapping from the YAML definition."""
        mapping = _load_sheet_mapping()
        result: list[dict] = []

        # BC NGÀY sheet column layout (34 columns, 0-indexed A-AH)
        # Matches excel_kv30_reader._COL_* constants
        col_headers = [
            "NGÀY", "THÁNG",
            "VỤ CHÁY THỐNG KÊ", "SCLQ ĐẾN PCCC&CNCH", "CHI VIỆN", "CNCH",
            "ĐỊNH KỲ NHÓM I", "ĐỊNH KỲ NHÓM II",
            "ĐỘT XUẤT NHÓM I", "ĐỘT XUẤT NHÓM II",
            "HƯỚNG DẪN", "KIẾN NGHỊ",
            "XỬ PHẠT", "TIỀN PHẠT", "ĐÌNH CHỈ", "PHỤC HỒI",
            "TIN BÀI", "PHÓNG SỰ", "SỐ LỚP TUYÊN TRUYỀN", "SỐ NGƯỜI THAM DỰ TUYÊN TRUYỀN",
            "SỐ KHUYẾN CÁO", "SỐ LỚP HUẤN LUYỆN", "SỐ NGƯỜI THAM DỰ HUẤN LUYỆN",
            "TỔNG SỐ LỚP", "TỔNG SỐ NGƯỜI",
            "PA PC06 XÂY DỰNG", "PA PC06 THỰC TẬP",
            "PA PC08 XÂY DỰNG", "PA PC08 THỰC TẬP",
            "PA PC09 XÂY DỰNG", "PA PC09 THỰC TẬP",
            "PA PC07 XÂY DỰNG", "PA PC07 THỰC TẬP",
            "Ghi chú",
        ]

        # STT field for each column (from excel_kv30_reader and sheet_mapping.yaml)
        col_stt_fields = [
            None,            # A: NGÀY (metadata)
            None,            # B: THÁNG (metadata)
            "stt_02_tong_chay",       # C
            "stt_14_tong_cnch",       # D
            "tong_chi_vien",           # E
            "stt_14_tong_cnch",        # F: CNCH supplement
            "stt_32_kiem_tra_dinh_ky", # G
            "stt_32_kiem_tra_dinh_ky", # H
            "stt_33_kiem_tra_dot_xuat", # I
            "stt_33_kiem_tra_dot_xuat", # J
            None,            # K: Hướng dẫn
            None,            # L: Kiến nghị
            "stt_35_xu_phat_tong",     # M
            "stt_40_xu_phat_tien",     # N
            "stt_38_xu_phat_dinh_chi", # O
            "stt_37_xu_phat_tam_dinh_chi", # P
            "stt_22_tt_mxh_tong",      # Q
            "stt_22_tt_mxh_tong",      # R: phóng sự → same STT
            "stt_29_tt_to_roi",        # S
            "stt_28_tt_so_nguoi",      # T
            "stt_29_tt_to_roi",        # U: khuyến cáo → same
            "stt_55_hl_tong_cbcs",     # V
            "stt_55_hl_tong_cbcs",     # W
            "stt_55_hl_tong_cbcs",     # X
            "stt_55_hl_tong_cbcs",     # Y
            "stt_43_pa_co_so_duyet",   # Z
            "stt_44_pa_co_so_thuc_tap", # AA
            "stt_46_pa_giao_thong_duyet", # AB
            "stt_47_pa_giao_thong_thuc_tap", # AC
            "stt_49_pa_cong_an_duyet",  # AD
            "stt_50_pa_cong_an_thuc_tap", # AE
            "stt_52_pa_cnch_ca_duyet",  # AF
            "stt_53_pa_cnch_ca_thuc_tap", # AG
            None,            # AH: Ghi chú
        ]

        btk_section = _as_dict(mapping).get("sheet_mapping", {}).get("bang_thong_ke", {})
        stt_map = _as_dict(btk_section).get("stt_map", {})

        for idx, (header, field) in enumerate(zip(col_headers, col_stt_fields)):
            stt_num: Optional[str] = None
            if field and field.startswith("stt_"):
                # Extract STT number from field name like "stt_02_tong_chay"
                parts = field.split("_")
                if len(parts) >= 2:
                    stt_num = parts[1]

            yaml_info = stt_map.get(stt_num, {}) if stt_num else {}
            yaml_field = yaml_info.get("field") if isinstance(yaml_info, dict) else None

            if field or yaml_field:
                status = "mapped" if (field or yaml_field) else "unmapped"
            elif header == "Ghi chú":
                status = "skipped"
            else:
                status = "unmapped"

            result.append({
                "col_index": idx,
                "col_letter": chr(65 + idx) if idx < 26 else chr(65 + idx // 26 - 1) + chr(65 + idx % 26),
                "col_header": header,
                "stt": stt_num,
                "field": field or yaml_field or "",
                "status": status,
            })

        return result

    # ── 4. Sheet names ────────────────────────────────────────────────────────

    def get_sheet_names(self, document_id: uuid.UUID) -> dict:
        """Download Excel from MinIO and return sheet names."""
        doc = self.db.query(Document).filter(Document.id == document_id).first()
        if not doc:
            return {"sheets": [], "error": "Document not found"}

        excel_bytes = _get_excel_bytes(doc)
        if not excel_bytes:
            return {"sheets": [], "error": "Could not download file from MinIO"}

        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            return {"sheets": wb.sheetnames}
        except Exception as e:
            return {"sheets": [], "error": str(e)}
